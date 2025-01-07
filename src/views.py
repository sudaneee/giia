# src/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import *
from datetime import datetime, date
import openpyxl

from openpyxl.styles import Alignment
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from django.utils.dateparse import parse_date
from django.contrib.auth.models import User
import os
from django.conf import settings
from django.utils import timezone
from datetime import timedelta
from django.utils.timezone import now
import logging
from django.db import transaction
from django.db.models import Sum, Q, F

import pandas as pd
from io import BytesIO
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required

import io
import xlsxwriter

def ordinal(n):
    """Convert an integer into its ordinal representation."""
    if 10 <= n % 100 <= 20:
        suffix = 'th'
    else:
        suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(n % 10, 'th')
    return str(n) + suffix



def user_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            request.session['username'] = username  # Set session data
            return redirect('student_list')
        else:
            messages.error(request, 'Invalid username or password')
    return render(request, 'src/login.html')


@login_required(login_url='login')
def user_logout(request):
    logout(request)
    request.session.flush()  # Clear all session data
    return redirect('login')


# View to list all students with bulk upload and class filter
@login_required(login_url='login')
def student_list(request):
    # Get the class filter from the query parameters
    class_id = request.GET.get('class_id', None)
    
    # Filter students based on the selected class
    if class_id:
        students = Student.objects.filter(admission_status='admitted', enrolled_class_id=class_id).all()
    else:
        students = Student.objects.filter(admission_status='admitted').all()
    
    # Get all available school classes for dropdowns
    school_classes = SchoolClass.objects.all()
    
    return render(request, 'src/student_list.html', {
        'students': students,
        'school_classes': school_classes,
        'selected_class_id': class_id  # Pass the selected class ID to the template
    })

# src/views.py
@login_required(login_url='login')
def add_student(request):
    if request.method == 'POST':
        # Get student form data
        first_name = request.POST.get('first_name', '')
        last_name = request.POST.get('last_name', '')
        date_of_birth = request.POST.get('date_of_birth', '')
        gender = request.POST.get('gender', '')
        address = request.POST.get('address', '')
        phone_number = request.POST.get('phone_number', '')
        email = request.POST.get('email', '')
        enrolled_class_id = request.POST.get('enrolled_class')
        status = request.POST.get('status', 'inactive')
        admission_number = request.POST.get('admission_number', '')

        # Get guardian form data
        guardian_first_name = request.POST.get('guardian_first_name', '')
        guardian_last_name = request.POST.get('guardian_last_name', '')
        guardian_phone_number = request.POST.get('guardian_phone_number', '')
        guardian_email = request.POST.get('guardian_email', '')
        guardian_relationship = request.POST.get('guardian_relationship', '')

        # Get the enrolled class if provided
        enrolled_class = get_object_or_404(SchoolClass, id=enrolled_class_id) if enrolled_class_id else None

        # Create the guardian object
        guardian = Guardian(
            first_name=guardian_first_name,
            last_name=guardian_last_name,
            phone_number=guardian_phone_number,
            email=guardian_email,
            relationship=guardian_relationship
        )
        guardian.save()

        # Create and save the new student
        student = Student(
            first_name=first_name,
            last_name=last_name,
            date_of_birth=date_of_birth,
            gender=gender,
            address=address,
            phone_number=phone_number,
            email=email,
            enrolled_class=enrolled_class,
            status=status,
        )
        student.save()

        # Link the student with the guardian
        student.guardians.add(guardian)
        student.save()

        # Generate admission number if not provided
        if not admission_number:
            admission_number = f"GIIA-{datetime.now().year}-{student.id}"
            student.admission_number = admission_number
            student.save()

        messages.success(request, 'Student and guardian added successfully!')
        return redirect('student_list')

    # If GET request, render the add student form
    school_classes = SchoolClass.objects.all()
    return render(request, 'src/add_student.html', {'school_classes': school_classes})




@login_required(login_url='login')
def update_student(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    guardians = student.guardians.all()

    if request.method == 'POST':
        # Get student form data
        student.first_name = request.POST.get('first_name', '')
        student.last_name = request.POST.get('last_name', '')
        student.date_of_birth = request.POST.get('date_of_birth', '')
        student.gender = request.POST.get('gender', '')
        student.address = request.POST.get('address', '')
        student.phone_number = request.POST.get('phone_number', '')
        student.email = request.POST.get('email', '')
        enrolled_class_id = request.POST.get('enrolled_class')
        student.status = request.POST.get('status', 'inactive')
        student.admission_number = request.POST.get('admission_number', student.admission_number)

        # Get or create the enrolled class
        student.enrolled_class = get_object_or_404(SchoolClass, id=enrolled_class_id) if enrolled_class_id else None

        # Update or create guardian details
        guardian_first_name = request.POST.get('guardian_first_name', '')
        guardian_last_name = request.POST.get('guardian_last_name', '')
        guardian_phone_number = request.POST.get('guardian_phone_number', '')
        guardian_email = request.POST.get('guardian_email', '')
        guardian_relationship = request.POST.get('guardian_relationship', '')

        # Assuming only one guardian per student; extend logic if there are multiple
        if guardians.exists():
            guardian = guardians.first()  # Assume updating the first guardian
            guardian.first_name = guardian_first_name
            guardian.last_name = guardian_last_name
            guardian.phone_number = guardian_phone_number
            guardian.email = guardian_email
            guardian.relationship = guardian_relationship
            guardian.save()
        else:
            # Create new guardian if none exist
            guardian = Guardian(
                first_name=guardian_first_name,
                last_name=guardian_last_name,
                phone_number=guardian_phone_number,
                email=guardian_email,
                relationship=guardian_relationship
            )
            guardian.save()
            student.guardians.add(guardian)

        # Save updated student data
        student.save()

        messages.success(request, 'Student and guardian details updated successfully!')
        return redirect('student_list')

    # If GET request, render the update student form
    school_classes = SchoolClass.objects.all()
    context = {
        'student': student,
        'guardians': guardians,
        'school_classes': school_classes,
    }
    return render(request, 'src/update_student.html', context)



# View to delete a student
@login_required(login_url='login')
def delete_student(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    student.delete()
    messages.success(request, 'Student deleted successfully!')
    return redirect('student_list')


# src/views.py



# Existing views (student_list, add_student, update_student, delete_student, bulk_upload_students)
# src/views.py

# src/views.py
@login_required(login_url='login')
def download_excel_template(request):
    # Create an Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Upload Template"

    # Define headers
    headers = [
        "First Name", "Last Name", "Date of Birth (YYYY-MM-DD)", "Gender", "Address",
        "Phone Number", "Email", "Class Name", "Status",
        "Guardian First Name", "Guardian Last Name",
        "Guardian Phone Number", "Guardian Email", "Guardian Relationship"
    ]
    ws.append(headers)

    # Set column widths to accommodate the headers
    column_widths = {
        'A': 15,  # First Name
        'B': 15,  # Last Name
        'C': 20,  # Date of Birth
        'D': 30,  # Address
        'E': 15,  # Phone Number
        'F': 25,  # Email
        'G': 20,  # Class Name
        'H': 10,  # Status
        'I': 18,  # Guardian First Name
        'J': 18,  # Guardian Last Name
        'K': 20,  # Guardian Phone Number
        'L': 25,  # Guardian Email
        'M': 18,   # Guardian Relationship
        'N': 18,
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # No pre-population with existing data to keep the template blank for new uploads

    # Prepare the response to download the file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=student_upload_template.xlsx'
    wb.save(response)

    return response


@login_required(login_url='login')
def bulk_upload_students(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        class_id = request.POST['enrolled_class']
        enrolled_class = SchoolClass.objects.get(id=class_id)
        excel_file = request.FILES['excel_file']

        # Check file typead
        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, 'Please upload a valid Excel file (.xlsx)')
            return redirect('student_list')

        # Read the Excel file
        try:
            wb = openpyxl.load_workbook(excel_file)
            worksheet = wb.active
        except Exception as e:
            messages.error(request, f'Error reading Excel file: {str(e)}')
            return redirect('student_list')

        # Iterate over the rows in the Excel file
        for row in worksheet.iter_rows(min_row=2, values_only=True):  # Skipping the header row
            first_name, last_name, dob_str, gender, = row

            # Parse date of birth
            if isinstance(dob_str, str):
                date_of_birth = parse_date(dob_str)
            elif isinstance(dob_str, date):  # Correctly reference date type from datetime import
                date_of_birth = dob_str
            else:
                date_of_birth = None

            # Get or create the enrolled class
            # if gender == 'Male':
            #     class_name = class_name + ' BOYS'
            # else:
            #     class_name = class_name + ' GIRLS'

            # enrolled_class = SchoolClass.objects.filter(name=class_name).first()

            # Create or update guardian
            # guardian = None
            # if guardian_first_name and guardian_last_name:
            #     guardian, created = Guardian.objects.get_or_create(
            #         first_name=guardian_first_name,
            #         last_name=guardian_last_name,
            #         phone_number=guardian_phone_number,
            #         email=guardian_email,
            #         defaults={'relationship': guardian_relationship}
            #     )

            # Create or update the student
            student = Student.objects.create(
                first_name=first_name,
                last_name=last_name,
                date_of_birth=date_of_birth,
                gender=gender,
                enrolled_class=enrolled_class,

            )

            # Generate admission number in the required format if not already set
            if not student.admission_number:
                if student.admission_status == "admitted":
                    student.admission_number = f"GIIA-{datetime.now().year}-{student.id}"
                else:
                    student.admission_number = f"APP-{datetime.now().year}-{student.id}"
                student.save()

            # if not created:
            #     # Update existing student details
            #     student.address = address
            #     student.phone_number = phone_number
            #     student.email = email
            #     student.enrolled_class = enrolled_class
            #     student.status = status
            #     student.save()

            # # Link guardian to student
            # if guardian:
            #     student.guardians.add(guardian)
            #     student.save()

        messages.success(request, 'Students uploaded successfully!')
        return redirect('student_list')



    return render(request, 'src/student_list.html')






logger = logging.getLogger(__name__)

@login_required(login_url='login')
def not_admitted_students(request):
    not_admitted_students = Student.objects.filter(admission_status='not_admitted')

    if request.method == 'POST':
        print("Request POST data:", request.POST)
        print("Request FILES data:", request.FILES)

        if 'bulk_admit' in request.POST:  # Bulk admission
            excel_file = request.FILES.get('excel_file', None)
            if not excel_file:
                messages.error(request, 'No Excel file uploaded.')
                return redirect('not_admitted_students')

            if not excel_file.name.endswith('.xlsx'):
                messages.error(request, 'Please upload a valid Excel file (.xlsx)')
                return redirect('not_admitted_students')

            try:
                wb = openpyxl.load_workbook(excel_file)
                worksheet = wb.active

                for row in worksheet.iter_rows(min_row=2, values_only=True):  # Skipping the header row
                    admission_number = row[0]
                    class_name = row[2]  # Assuming class name is in the third column

                    try:
                        student = Student.objects.get(admission_number=admission_number)
                    except Student.DoesNotExist:
                        messages.error(request, f'Student with admission number {admission_number} not found.')
                        continue

                    if class_name:
                        enrolled_class, created = SchoolClass.objects.get_or_create(name=class_name, arm="A")
                        student.enrolled_class = enrolled_class
                    else:
                        messages.error(request, f'No class specified for admission number {admission_number}.')
                        continue

                    student.admission_status = 'admitted'
                    student.admitted_at = datetime.now()
                    student.admission_number = f"GIIA-{datetime.now().year}-{student.id}"
                    student.save()

                messages.success(request, 'Students have been admitted successfully using the Excel file.')
            except Exception as e:
                messages.error(request, f'Error processing the Excel file: {str(e)}')
            return redirect('not_admitted_students')

        elif 'admit_selected' in request.POST:  # Admit selected students
            student_ids = request.POST.getlist('student_ids')

            admitted_count = 0
            for student_id in student_ids:
                try:
                    student = Student.objects.get(id=student_id)
                    if student.enrolled_class is None:
                        messages.error(request, f'Student {student.first_name} {student.last_name} has no class assigned. Cannot admit without a class.')
                        continue

                    student.admission_status = 'admitted'
                    student.admission_number = f"GIIA-{datetime.now().year}-{student.id}"
                    student.admitted_at = datetime.now()
                    student.save()
                    admitted_count += 1
                except Student.DoesNotExist:
                    messages.error(request, f'Student with ID {student_id} not found.')
                    continue

            if admitted_count > 0:
                messages.success(request, f'{admitted_count} selected students have been admitted successfully!')
            else:
                messages.error(request, 'No students were admitted. Please check your selection.')

            return redirect('not_admitted_students')

    # Default GET request rendering
    return render(request, 'src/not_admitted_students.html', {'students': not_admitted_students})




# src/views.py

@login_required(login_url='login')
def download_not_admitted_template(request):
    # Create an Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Not Admitted Students Template"

    # Define headers
    headers = ["Admission Number", "Name", "Class Name"]
    ws.append(headers)

    # Filter not admitted students and populate the Excel file
    not_admitted_students = Student.objects.filter(admission_status='not_admitted')
    for student in not_admitted_students:
        ws.append([
            student.admission_number,
            f"{student.first_name} {student.last_name}",
            student.enrolled_class.name if student.enrolled_class else ""  # Include class name if assigned
        ])

    # Prepare the response to download the file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=not_admitted_students_template.xlsx'
    wb.save(response)

    return response







def generate_admission_letter(request, student_id):
    # Fetch student details
    student = get_object_or_404(Student, id=student_id)

    # Fetch school configuration (assuming only one configuration is needed)
    school_config = SchoolConfig.objects.first()

    if not school_config:
        messages.error(request, 'School configuration is missing. Please upload header and signature images.')
        return redirect('student_list')

    print(f"Student admission status: {student.admission_status}")

    if student.admission_status == 'admitted':
        # Create a response object and set the content type to PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{student.first_name}_{student.last_name}_admission_letter.pdf"'

        # Create a canvas
        p = canvas.Canvas(response, pagesize=A4)
        width, height = A4

        # Paths to images
        header_image_path = os.path.join(settings.MEDIA_ROOT, school_config.header_image.name)
        signature_image_path = os.path.join(settings.MEDIA_ROOT, school_config.signature_image.name)

        # Debugging: Log the paths
        print(f"Header Image Path: {header_image_path}")
        print(f"Signature Image Path: {signature_image_path}")
        print(f"Header image exists: {os.path.exists(header_image_path)}")
        print(f"Signature image exists: {os.path.exists(signature_image_path)}")

        # Check if the image paths exist before proceeding
        if not os.path.exists(header_image_path):
            messages.error(request, 'Header image file is missing in media folder.')
            return redirect('student_list')

        if not os.path.exists(signature_image_path):
            messages.error(request, 'Signature image file is missing in media folder.')
            return redirect('student_list')

        # Add the school header image
        p.drawImage(header_image_path, x=1 * inch, y=height - 2.5 * inch, width=width - 2 * inch, height=2 * inch)

        # Add Date, Admission No, Session
        p.setFont("Helvetica-Bold", 12)
        p.drawString(1 * inch, height - 3.4 * inch, f"Date: {student.admitted_at.strftime('%Y-%m-%d')}")

        p.drawString(1 * inch, height - 3.6 * inch, f"Admission No: {student.admission_number}")
        p.drawString(1 * inch, height - 3.8 * inch, f"Admitted Class: {student.enrolled_class}")
        p.drawString(1 * inch, height - 4.0 * inch, f"Session: 2024/2025")

        # Add Provisional Admission Letter title
        p.setFont("Helvetica-Bold", 14)
        p.drawString(1 * inch, height - 4.2 * inch, "PROVISIONAL ADMISSION LETTER")

        # Add the student's name
        p.setFont("Helvetica", 12)
        p.drawString(1 * inch, height - 4.6 * inch, f"Name: {student.first_name} {student.last_name}")

        # Add the body of the letter
        p.drawString(1 * inch, height - 5.0 * inch, "1. We are pleased to inform you that, due to your success in the interview,")
        p.drawString(1 * inch, height - 5.2 * inch, "   you have been offered a provisional admission into the Great Insight International")
        p.drawString(1 * inch, height - 5.4 * inch, "   Academy, Zaria.")
        p.drawString(1 * inch, height - 5.8 * inch, "2. You are expected to report to the school on or before two weeks from the date of admission")
        p.drawString(1 * inch, height - 6.0 * inch, "   accompanied by the following:")
        p.drawString(1.2 * inch, height - 6.2 * inch, "i. evidence of payment")
        p.drawString(1.2 * inch, height - 6.4 * inch, "ii. a copy of admission letter")
        p.drawString(1.2 * inch, height - 6.6 * inch, "iii. birth certificate")
        p.drawString(1.2 * inch, height - 6.8 * inch, "iv. two recent passport sized photographs")
        p.drawString(1.2 * inch, height - 7.0 * inch, "v. a copy of blood group and genotype test results from a recognized government hospital")
        p.drawString(1.2 * inch, height - 7.2 * inch, "vi. a copy of transfer letter/evidence of last term school fees from previous school attended")

        p.drawString(1 * inch, height - 7.4 * inch, "Failure to comply will result in forfeiture of your admission.")

        # Add the signature image
        p.drawImage(signature_image_path, x=1 * inch, y=height - 8.5 * inch, width=2 * inch, height=1 * inch)

        # Signature text below the signature image
        p.drawString(1 * inch, height - 9.0 * inch, "Ustz. Aliyu Ibrahim Yerima")
        p.drawString(1 * inch, height - 9.2 * inch, "Head Teacher")

        # Save and return the PDF
        p.showPage()
        p.save()

        return response

    messages.error(request, 'Student not admitted yet!')
    return redirect('student_list')






def generate_admission_letter(request, student_id):
    # Fetch student details
    student = get_object_or_404(Student, id=student_id)

    # Fetch school configuration (assuming only one configuration is needed)
    school_config = SchoolConfig.objects.first()

    if not school_config:
        messages.error(request, 'School configuration is missing. Please upload header and signature images.')
        return redirect('student_list')

    print(f"Student admission status: {student.admission_status}")

    if student.admission_status == 'admitted':
        # Create a response object and set the content type to PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{student.first_name}_{student.last_name}_admission_letter.pdf"'

        # Create a canvas
        p = canvas.Canvas(response, pagesize=A4)
        width, height = A4

        # Paths to images
        header_image_path = os.path.join(settings.MEDIA_ROOT, school_config.header_image.name)
        signature_image_path = os.path.join(settings.MEDIA_ROOT, school_config.signature_image.name)

        # Debugging: Log the paths
        print(f"Header Image Path: {header_image_path}")
        print(f"Signature Image Path: {signature_image_path}")
        print(f"Header image exists: {os.path.exists(header_image_path)}")
        print(f"Signature image exists: {os.path.exists(signature_image_path)}")

        # Check if the image paths exist before proceeding
        if not os.path.exists(header_image_path):
            messages.error(request, 'Header image file is missing in media folder.')
            return redirect('student_list')

        if not os.path.exists(signature_image_path):
            messages.error(request, 'Signature image file is missing in media folder.')
            return redirect('student_list')

        # Add the school header image
        p.drawImage(header_image_path, x=1 * inch, y=height - 2.5 * inch, width=width - 2 * inch, height=2 * inch)

        # Add Date, Admission No, Session
        p.setFont("Helvetica-Bold", 12)
        p.drawString(1 * inch, height - 3.4 * inch, f"Date: {student.admitted_at.strftime('%Y-%m-%d')}")

        p.drawString(1 * inch, height - 3.6 * inch, f"Admission No: {student.admission_number}")
        p.drawString(1 * inch, height - 3.8 * inch, f"Admitted Class: {student.enrolled_class}")
        p.drawString(1 * inch, height - 4.0 * inch, f"Session: 2024/2025")

        # Add Provisional Admission Letter title
        p.setFont("Helvetica-Bold", 14)
        p.drawString(1 * inch, height - 4.2 * inch, "PROVISIONAL ADMISSION LETTER")

        # Add the student's name
        p.setFont("Helvetica", 12)
        p.drawString(1 * inch, height - 4.6 * inch, f"Name: {student.first_name} {student.last_name}")

        # Add the body of the letter
        p.drawString(1 * inch, height - 5.0 * inch, "1. We are pleased to inform you that, due to your success in the interview,")
        p.drawString(1 * inch, height - 5.2 * inch, "   you have been offered a provisional admission into the Great Insight International")
        p.drawString(1 * inch, height - 5.4 * inch, "   Academy, Zaria.")
        p.drawString(1 * inch, height - 5.8 * inch, "2. You are expected to report to the school on or before two weeks from the date of admission")
        p.drawString(1 * inch, height - 6.0 * inch, "   accompanied by the following:")
        p.drawString(1.2 * inch, height - 6.2 * inch, "i. evidence of payment")
        p.drawString(1.2 * inch, height - 6.4 * inch, "ii. a copy of admission letter")
        p.drawString(1.2 * inch, height - 6.6 * inch, "iii. birth certificate")
        p.drawString(1.2 * inch, height - 6.8 * inch, "iv. two recent passport sized photographs")
        p.drawString(1.2 * inch, height - 7.0 * inch, "v. a copy of blood group and genotype test results from a recognized government hospital")
        p.drawString(1.2 * inch, height - 7.2 * inch, "vi. a copy of transfer letter/evidence of last term school fees from previous school attended")

        p.drawString(1 * inch, height - 7.4 * inch, "Failure to comply will result in forfeiture of your admission.")

        # Add the signature image
        p.drawImage(signature_image_path, x=1 * inch, y=height - 8.5 * inch, width=2 * inch, height=1 * inch)

        # Signature text below the signature image
        p.drawString(1 * inch, height - 9.0 * inch, "Ustz. Aliyu Ibrahim Yerima")
        p.drawString(1 * inch, height - 9.2 * inch, "Head Teacher")

        # Save and return the PDF
        p.showPage()
        p.save()

        return response

    messages.error(request, 'Student not admitted yet!')
    return redirect('student_list')



def admitted_students(request):
    # admitted_students = Student.objects.filter(admission_status='admitted')
    
    # Get the start of today
    start_of_today = now().replace(hour=0, minute=0, second=0, microsecond=0)
    
    # Calculate the date range: start of today and two weeks back
    two_weeks_ago = start_of_today - timedelta(weeks=2)
    
    Filter admitted students within the date range
    admitted_students = Student.objects.filter(
        admission_status='admitted',
        created_at__gte=two_weeks_ago
    )
    not_admitted_students = Student.objects.filter(admission_status='not_admitted')

    if request.method == 'POST':
        # Check if the form is submitted for generating the admission letter
        if 'generate_admission_letter' in request.POST:
            student_id = request.POST.get('student_id')
            source = request.POST.get('source')

            try:
                student = Student.objects.get(id=student_id)

                # Check if feedback already exists
                how_you_find_us, created = HowYouFindUs.objects.get_or_create(student=student)
                if not created:
                    # If feedback already exists, skip updating and proceed to admission letter generation
                    messages.info(request, 'You have already submitted feedback. Generating admission letter.')
                    return redirect('generate_admission_letter', student_id=student_id)

                # If no feedback exists, update with new source information
                how_you_find_us.source = source
                how_you_find_us.save()

                # Redirect to generate the admission letter
                return redirect('generate_admission_letter', student_id=student_id)

            except Student.DoesNotExist:
                messages.error(request, 'Student not found.')

    return render(request, 'src/admitted_students.html', {
        'admitted_students': admitted_students,
        'not_admitted_students': not_admitted_students,
    })





@login_required(login_url='login')
def subject_list(request):
    subjects = Subject.objects.all()
    return render(request, 'src/subject_list.html', {'subjects': subjects})
@login_required(login_url='login')
def subject_create(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')
        Subject.objects.create(name=name, description=description)
        return redirect('subject_list')
    return render(request, 'src/subject_form.html')
@login_required(login_url='login')
def subject_update(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    if request.method == 'POST':
        subject.name = request.POST.get('name')
        subject.description = request.POST.get('description')
        subject.save()
        return redirect('subject_list')
    return render(request, 'src/subject_form.html', {'subject': subject})
@login_required(login_url='login')
def subject_delete(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    if request.method == 'POST':
        subject.delete()
        return redirect('subject_list')
    return render(request, 'src/subject_confirm_delete.html', {'subject': subject})

# Session CRUD
@login_required(login_url='login')
def session_list(request):
    sessions = Session.objects.all()
    return render(request, 'src/session_list.html', {'sessions': sessions})
@login_required(login_url='login')
def session_create(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        current = request.POST.get('current') == 'on'
        Session.objects.create(name=name, start_date=start_date, end_date=end_date, current=current)
        return redirect('session_list')
    return render(request, 'src/session_form.html')
@login_required(login_url='login')
def session_update(request, pk):
    session = get_object_or_404(Session, pk=pk)
    if request.method == 'POST':
        session.name = request.POST.get('name')
        session.start_date = request.POST.get('start_date')
        session.end_date = request.POST.get('end_date')
        session.current = request.POST.get('current') == 'on'
        session.save()
        return redirect('session_list')
    return render(request, 'src/session_form.html', {'session': session})
@login_required(login_url='login')
def session_delete(request, pk):
    session = get_object_or_404(Session, pk=pk)
    if request.method == 'POST':
        session.delete()
        return redirect('session_list')
    return render(request, 'src/session_confirm_delete.html', {'session': session})

# Term CRUD
@login_required(login_url='login')
def term_list(request):
    terms = Term.objects.all()
    return render(request, 'src/term_list.html', {'terms': terms})
@login_required(login_url='login')
def term_create(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        session_id = request.POST.get('session')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        session = get_object_or_404(Session, pk=session_id)
        Term.objects.create(name=name, session=session, start_date=start_date, end_date=end_date)
        return redirect('term_list')
    sessions = Session.objects.all()
    return render(request, 'src/term_form.html', {'sessions': sessions})
@login_required(login_url='login')
def term_update(request, pk):
    term = get_object_or_404(Term, pk=pk)
    if request.method == 'POST':
        term.name = request.POST.get('name')
        session_id = request.POST.get('session')
        term.start_date = request.POST.get('start_date')
        term.end_date = request.POST.get('end_date')
        term.session = get_object_or_404(Session, pk=session_id)
        term.save()
        return redirect('term_list')
    sessions = Session.objects.all()
    return render(request, 'src/term_form.html', {'term': term, 'sessions': sessions})
@login_required(login_url='login')
def term_delete(request, pk):
    term = get_object_or_404(Term, pk=pk)
    if request.method == 'POST':
        term.delete()
        return redirect('term_list')
    return render(request, 'src/term_confirm_delete.html', {'term': term})

# SchoolClass CRUD
@login_required(login_url='login')
def schoolclass_list(request):
    classes = SchoolClass.objects.all()
    return render(request, 'src/schoolclass_list.html', {'classes': classes})
@login_required(login_url='login')
def schoolclass_create(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')
        level = request.POST.get('level')
        arm = request.POST.get('arm')
        SchoolClass.objects.create(name=name, description=description, level=level, arm=arm)
        return redirect('schoolclass_list')
    return render(request, 'src/schoolclass_form.html')
@login_required(login_url='login')
def schoolclass_update(request, pk):
    schoolclass = get_object_or_404(SchoolClass, pk=pk)
    if request.method == 'POST':
        schoolclass.name = request.POST.get('name')
        schoolclass.description = request.POST.get('description')
        schoolclass.level = request.POST.get('level')
        schoolclass.arm = request.POST.get('arm')
        schoolclass.save()
        return redirect('schoolclass_list')
    return render(request, 'src/schoolclass_form.html', {'schoolclass': schoolclass})
@login_required(login_url='login')
def schoolclass_delete(request, pk):
    schoolclass = get_object_or_404(SchoolClass, pk=pk)
    if request.method == 'POST':
        schoolclass.delete()
        return redirect('schoolclass_list')
    return render(request, 'src/schoolclass_confirm_delete.html', {'schoolclass': schoolclass})

# FeeStructure CRUD
@login_required(login_url='login')
def feestructure_list(request):
    feestructures = FeeStructure.objects.all()
    return render(request, 'src/feestructure_list.html', {'feestructures': feestructures})
@login_required(login_url='login')
def feestructure_create(request):
    if request.method == 'POST':
        class_id = request.POST.get('class_assigned')
        description = request.POST.get('description')
        amount = request.POST.get('amount')
        due_date = request.POST.get('due_date')
        session_id = request.POST.get('session')
        term_id = request.POST.get('term')
        fee_type = request.POST.get('fee_type')
        class_assigned = get_object_or_404(SchoolClass, pk=class_id)
        session = get_object_or_404(Session, pk=session_id)
        term = get_object_or_404(Term, pk=term_id)
        FeeStructure.objects.create(
            class_assigned=class_assigned,
            description=description,
            amount=amount,
            due_date=due_date,
            session=session,
            term=term,
            fee_type=fee_type
        )
        return redirect('feestructure_list')
    classes = SchoolClass.objects.all()
    sessions = Session.objects.all()
    terms = Term.objects.all()
    return render(request, 'src/feestructure_form.html', {'classes': classes, 'sessions': sessions, 'terms': terms})
@login_required(login_url='login')
def feestructure_update(request, pk):
    feestructure = get_object_or_404(FeeStructure, pk=pk)
    if request.method == 'POST':
        class_id = request.POST.get('class_assigned')
        feestructure.description = request.POST.get('description')
        feestructure.amount = request.POST.get('amount')
        feestructure.due_date = request.POST.get('due_date')
        session_id = request.POST.get('session')
        term_id = request.POST.get('term')
        feestructure.fee_type = request.POST.get('fee_type')
        feestructure.class_assigned = get_object_or_404(SchoolClass, pk=class_id)
        feestructure.session = get_object_or_404(Session, pk=session_id)
        feestructure.term = get_object_or_404(Term, pk=term_id)
        feestructure.save()
        return redirect('feestructure_list')
    classes = SchoolClass.objects.all()
    sessions = Session.objects.all()
    terms = Term.objects.all()
    return render(request, 'src/feestructure_form.html', {'feestructure': feestructure, 'classes': classes, 'sessions': sessions, 'terms': terms})
@login_required(login_url='login')
def feestructure_delete(request, pk):
    feestructure = get_object_or_404(FeeStructure, pk=pk)
    if request.method == 'POST':
        feestructure.delete()
        return redirect('feestructure_list')
    return render(request, 'src/feestructure_confirm_delete.html', {'feestructure': feestructure})

# Payment CRUD

@login_required(login_url='login')
def payment_list(request):
    payments = Payment.objects.all()
    students = Student.objects.all()
    schoolclasses = SchoolClass.objects.all()
    feestructures = FeeStructure.objects.all()
    sessions = Session.objects.all()

    # Filtering logic
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    student_id = request.GET.get('student')
    school_class_id = request.GET.get('school_class')
    fee_structure_id = request.GET.get('fee_structure')
    payment_method = request.GET.get('payment_method')
    status = request.GET.get('status')
    session_id = request.GET.get('session')

    if start_date and end_date:
        payments = payments.filter(payment_date__range=[start_date, end_date])
    if student_id:
        payments = payments.filter(student_id=student_id)
    if school_class_id:
        payments = payments.filter(student__schoolclass__id=school_class_id)
    if fee_structure_id:
        payments = payments.filter(fee_structure_id=fee_structure_id)
    if payment_method:
        payments = payments.filter(payment_method=payment_method)
    if status:
        payments = payments.filter(status=status)
    if session_id:
        payments = payments.filter(session_id=session_id)

    # Calculate the total sum of filtered payments
    total_sum = payments.aggregate(total=Sum('amount_paid'))['total'] or 0

    return render(request, 'src/payment_list.html', {
        'payments': payments,
        'students': students,
        'schoolclasses': schoolclasses,
        'feestructures': feestructures,
        'sessions': sessions,
        'total_sum': total_sum
    })

@login_required(login_url='login')
def payment_create(request):
    if request.method == 'POST':
        student_id = request.POST.get('student')
        fee_structure_id = request.POST.get('fee_structure')
        amount_paid = request.POST.get('amount_paid')
        payment_method = request.POST.get('payment_method')
        status = request.POST.get('status')
        session_id = request.POST.get('session')
        term_id = request.POST.get('term')
        fee_structure = get_object_or_404(FeeStructure, pk=fee_structure_id)
        session = get_object_or_404(Session, pk=session_id)
        term = get_object_or_404(Term, pk=term_id)
        student = get_object_or_404(Student, pk=student_id)
        Payment.objects.create(
            student=student,
            fee_structure=fee_structure,
            amount_paid=amount_paid,
            payment_method=payment_method,
            status=status,
            session=session,
            term=term
        )
        return redirect('payment_list')
    fee_structures = FeeStructure.objects.all()
    sessions = Session.objects.all()
    terms = Term.objects.all()
    students = Student.objects.all()
    return render(request, 'src/payment_form.html', {'fee_structures': fee_structures, 'sessions': sessions, 'terms': terms, 'students': students})

@login_required(login_url='login')
def payment_update(request, pk):
    payment = get_object_or_404(Payment, pk=pk)
    if request.method == 'POST':
        student_id = request.POST.get('student')
        fee_structure_id = request.POST.get('fee_structure')
        payment.amount_paid = request.POST.get('amount_paid')
        payment.payment_method = request.POST.get('payment_method')
        payment.status = request.POST.get('status')
        session_id = request.POST.get('session')
        term_id = request.POST.get('term')
        payment.fee_structure = get_object_or_404(FeeStructure, pk=fee_structure_id)
        payment.session = get_object_or_404(Session, pk=session_id)
        payment.term = get_object_or_404(Term, pk=term_id)
        payment.student = get_object_or_404(Student, pk=student_id)
        payment.save()
        return redirect('payment_list')  # Use the name of the URL pattern

    fee_structures = FeeStructure.objects.all()
    sessions = Session.objects.all()
    terms = Term.objects.all()
    students = Student.objects.all()
    return render(request, 'src/payment_form.html', {
        'payment': payment,
        'fee_structures': fee_structures,
        'sessions': sessions,
        'terms': terms,
        'students': students
    })
@login_required(login_url='login')
def payment_delete(request, pk):
    payment = get_object_or_404(Payment, pk=pk)
    if request.method == 'POST':
        payment.delete()
        return redirect('payment_list')
    return render(request, 'src/payment_confirm_delete.html', {'payment': payment})

@login_required(login_url='login')
def payment_export_excel(request):
    payments = Payment.objects.all()

    # Apply filters (similar to `payment_list` view)
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    student_id = request.GET.get('student')
    school_class_id = request.GET.get('school_class')
    fee_structure_id = request.GET.get('fee_structure')
    payment_method = request.GET.get('payment_method')
    status = request.GET.get('status')
    session_id = request.GET.get('session')

    if start_date and end_date:
        payments = payments.filter(payment_date__range=[start_date, end_date])
    if student_id:
        payments = payments.filter(student_id=student_id)
    if school_class_id:
        payments = payments.filter(student__schoolclass__id=school_class_id)
    if fee_structure_id:
        payments = payments.filter(fee_structure_id=fee_structure_id)
    if payment_method:
        payments = payments.filter(payment_method=payment_method)
    if status:
        payments = payments.filter(status=status)
    if session_id:
        payments = payments.filter(session_id=session_id)

    # Calculate the total amount of filtered payments
    total_amount = payments.aggregate(total=Sum('amount_paid'))['total'] or 0

    # Create an Excel file
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Payments'

    # Write header
    headers = ['ID', 'Admission Number', 'Student', 'Amount Paid', 'Date', 'Status', 'Method', 'Term', 'Session']
    worksheet.append(headers)

    # Write data
    for payment in payments:
        worksheet.append([
            payment.id,
            payment.student.admission_number,  # Assuming 'admission_number' is a field on the Student model
            str(payment.student),
            payment.amount_paid,
            payment.payment_date.strftime('%Y-%m-%d'),
            payment.status,
            payment.payment_method,
            payment.term.name,
            payment.session.name,
        ])

    # Append the total amount at the bottom
    worksheet.append([])
    worksheet.append(['', '', 'Total', total_amount])

    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=filtered_payments.xlsx'
    workbook.save(response)

    return response




# Category Views
@login_required(login_url='login')
def category_list(request):
    categories = Category.objects.all()
    return render(request, 'src/category_list.html', {'categories': categories})

@login_required(login_url='login')
def category_create(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')
        Category.objects.create(name=name, description=description)
        return redirect('category_list')
    return render(request, 'src/category_form.html')


@login_required(login_url='login')
def category_update(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        category.name = request.POST.get('name')
        category.description = request.POST.get('description')
        category.save()
        return redirect('category_list')
    return render(request, 'src/category_form.html', {'category': category})


@login_required(login_url='login')
def category_delete(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        category.delete()
        return redirect('category_list')
    return render(request, 'src/category_confirm_delete.html', {'category': category})

# Supplier Views

@login_required(login_url='login')
def supplier_list(request):
    suppliers = Supplier.objects.all()
    return render(request, 'src/supplier_list.html', {'suppliers': suppliers})


@login_required(login_url='login')
def supplier_create(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        contact_email = request.POST.get('contact_email')
        contact_phone = request.POST.get('contact_phone')
        address = request.POST.get('address')
        Supplier.objects.create(name=name, contact_email=contact_email, contact_phone=contact_phone, address=address)
        return redirect('supplier_list')
    return render(request, 'src/supplier_form.html')


@login_required(login_url='login')
def supplier_update(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        supplier.name = request.POST.get('name')
        supplier.contact_email = request.POST.get('contact_email')
        supplier.contact_phone = request.POST.get('contact_phone')
        supplier.address = request.POST.get('address')
        supplier.save()
        return redirect('supplier_list')
    return render(request, 'src/supplier_form.html', {'supplier': supplier})


@login_required(login_url='login')
def supplier_delete(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        supplier.delete()
        return redirect('supplier_list')
    return render(request, 'src/supplier_confirm_delete.html', {'supplier': supplier})

# Item Views
@login_required(login_url='login')
def item_list(request):
    items = Item.objects.all()
    return render(request, 'src/item_list.html', {'items': items})


@login_required(login_url='login')
def item_create(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        category_id = request.POST.get('category')
        description = request.POST.get('description')
        quantity_in_stock = request.POST.get('quantity_in_stock')
        reorder_level = request.POST.get('reorder_level')
        expiry_date = request.POST.get('expiry_date')

        category = Category.objects.get(pk=category_id) if category_id else None

        Item.objects.create(name=name, category=category, description=description,
                            quantity_in_stock=quantity_in_stock, reorder_level=reorder_level,
                            expiry_date=expiry_date)
        return redirect('item_list')
    categories = Category.objects.all()
    return render(request, 'src/item_form.html', {'categories': categories})


@login_required(login_url='login')
def item_update(request, pk):
    item = get_object_or_404(Item, pk=pk)
    if request.method == 'POST':
        item.name = request.POST.get('name')
        category_id = request.POST.get('category')
        item.description = request.POST.get('description')
        item.quantity_in_stock = request.POST.get('quantity_in_stock')
        item.reorder_level = request.POST.get('reorder_level')
        item.expiry_date = request.POST.get('expiry_date')

        item.category = Category.objects.get(pk=category_id) if category_id else None
        item.save()
        return redirect('item_list')
    categories = Category.objects.all()
    return render(request, 'src/item_form.html', {'item': item, 'categories': categories})


@login_required(login_url='login')
def item_delete(request, pk):
    item = get_object_or_404(Item, pk=pk)
    if request.method == 'POST':
        item.delete()
        return redirect('item_list')
    return render(request, 'src/item_confirm_delete.html', {'item': item})

# InventoryTransaction Views
@login_required(login_url='login')
def transaction_list(request):
    transactions = InventoryTransaction.objects.all()
    return render(request, 'src/transaction_list.html', {'transactions': transactions})


@login_required(login_url='login')
def transaction_create(request):
    if request.method == 'POST':
        item_id = request.POST.get('item')
        transaction_type = request.POST.get('transaction_type')
        quantity = request.POST.get('quantity')
        description = request.POST.get('description')

        item = Item.objects.get(pk=item_id)
        InventoryTransaction.objects.create(item=item, transaction_type=transaction_type,
                                            quantity=quantity, description=description)
        return redirect('transaction_list')
    items = Item.objects.all()
    return render(request, 'src/transaction_form.html', {'items': items})


@login_required(login_url='login')
def transaction_update(request, pk):
    transaction = get_object_or_404(InventoryTransaction, pk=pk)
    if request.method == 'POST':
        transaction_type = request.POST.get('transaction_type')
        quantity = request.POST.get('quantity')
        description = request.POST.get('description')

        transaction.transaction_type = transaction_type
        transaction.quantity = quantity
        transaction.description = description
        transaction.save()
        return redirect('transaction_list')
    items = Item.objects.all()
    return render(request, 'src/transaction_form.html', {'transaction': transaction, 'items': items})


@login_required(login_url='login')
def transaction_delete(request, pk):
    transaction = get_object_or_404(InventoryTransaction, pk=pk)
    if request.method == 'POST':
        transaction.delete()
        return redirect('transaction_list')
    return render(request, 'src/transaction_confirm_delete.html', {'transaction': transaction})

# Purchase Order Views
@login_required(login_url='login')
def purchase_order_list(request):
    orders = PurchaseOrder.objects.all()
    return render(request, 'src/purchase_order_list.html', {'orders': orders})


@login_required(login_url='login')
def purchase_order_create(request):
    if request.method == 'POST':
        item_id = request.POST.get('item')
        quantity_ordered = request.POST.get('quantity_ordered')
        received_quantity = request.POST.get('received_quantity')
        price_per_unit = request.POST.get('price_per_unit')
        supplier_id = request.POST.get('supplier')
        received_date = request.POST.get('received_date')

        # Convert input strings to integers or floats
        quantity_ordered = int(quantity_ordered) if quantity_ordered else 0
        received_quantity = int(received_quantity) if received_quantity else 0
        price_per_unit = float(price_per_unit) if price_per_unit else 0.0

        item = get_object_or_404(Item, pk=item_id)
        supplier = get_object_or_404(Supplier, pk=supplier_id) if supplier_id else None

        # Create the PurchaseOrder instance
        order = PurchaseOrder.objects.create(
            item=item,
            quantity_ordered=quantity_ordered,
            received_quantity=received_quantity,
            price_per_unit=price_per_unit,
            supplier=supplier,
            received_date=received_date
        )

        # Update the item's quantity in stock
        item.quantity_in_stock += received_quantity
        item.save()

        return redirect('purchase_order_list')

    items = Item.objects.all()
    suppliers = Supplier.objects.all()
    return render(request, 'src/purchase_order_form.html', {'items': items, 'suppliers': suppliers})



@login_required(login_url='login')
def purchase_order_update(request, pk):
    order = get_object_or_404(PurchaseOrder, pk=pk)
    if request.method == 'POST':
        item_id = request.POST.get('item')
        quantity_ordered = request.POST.get('quantity_ordered')
        received_quantity = request.POST.get('received_quantity')
        price_per_unit = request.POST.get('price_per_unit')
        supplier_id = request.POST.get('supplier')
        received_date = request.POST.get('received_date')

        # Convert input strings to integers or floats
        quantity_ordered = int(quantity_ordered) if quantity_ordered else 0
        received_quantity = int(received_quantity) if received_quantity else 0
        price_per_unit = float(price_per_unit) if price_per_unit else 0.0

        item = get_object_or_404(Item, pk=item_id)
        supplier = get_object_or_404(Supplier, pk=supplier_id) if supplier_id else None

        # Update the PurchaseOrder instance
        order.item = item
        order.quantity_ordered = quantity_ordered
        order.received_quantity = received_quantity
        order.price_per_unit = price_per_unit
        order.supplier = supplier
        order.received_date = received_date
        order.save()

        # Update the item's quantity in stock based on the new received quantity
        item.quantity_in_stock += received_quantity
        item.save()

        return redirect('purchase_order_list')

    items = Item.objects.all()
    suppliers = Supplier.objects.all()
    return render(request, 'src/purchase_order_form.html', {'order': order, 'items': items, 'suppliers': suppliers})


@login_required(login_url='login')
def purchase_order_delete(request, pk):
    order = get_object_or_404(PurchaseOrder, pk=pk)
    if request.method == 'POST':
        order.delete()
        return redirect('purchase_order_list')
    return render(request, 'src/purchase_order_confirm_delete.html', {'order': order})





@login_required(login_url='login')
def result_entry(request):
    subjects = Subject.objects.all()
    classes = SchoolClass.objects.all()
    sessions = Session.objects.all()
    terms = Term.objects.all()

    if request.method == 'POST':
        term_id = request.POST.get('term')
        session_id = request.POST.get('session')
        class_id = request.POST.get('class_assigned')
        subject_id = request.POST.get('subject')

        if term_id and session_id and class_id and subject_id:
            term = get_object_or_404(Term, pk=term_id)
            session = get_object_or_404(Session, pk=session_id)
            school_class = get_object_or_404(SchoolClass, pk=class_id)
            subject = get_object_or_404(Subject, pk=subject_id)
            students = Student.objects.filter(enrolled_class=school_class)

            if 'save_results' in request.POST:
                for student in students:
                    # Ensure that marks are integers, defaulting to 0 if invalid or missing
                    ca1_marks = request.POST.get(f'ca1_{student.id}', '0')
                    ca2_marks = request.POST.get(f'ca2_{student.id}', '0')
                    home_work_marks = request.POST.get(f'home_work_{student.id}', '0')
                    activity_marks = request.POST.get(f'activity_{student.id}', '0')
                    exam_marks = request.POST.get(f'exam_{student.id}', '0')

                    # Use int() conversion directly and handle exceptions
                    try:
                        ca1_marks = int(ca1_marks)
                    except ValueError:
                        ca1_marks = 0

                    try:
                        ca2_marks = int(ca2_marks)
                    except ValueError:
                        ca2_marks = 0

                    try:
                        home_work_marks = int(home_work_marks)
                    except ValueError:
                        home_work_marks = 0

                    try:
                        activity_marks = int(activity_marks)
                    except ValueError:
                        activity_marks = 0

                    try:
                        exam_marks = int(exam_marks)
                    except ValueError:
                        exam_marks = 0

                    # Check if result already exists
                    if Result.objects.filter(
                        student=student,
                        subject=subject,
                        class_assigned=school_class,
                        session=session,
                        term=term
                    ).exists():
                        messages.warning(request, f"Result for {student.first_name} {student.last_name} already exists.")
                    else:
                        # Create a new result if it doesn't exist
                        Result.objects.create(
                            student=student,
                            subject=subject,
                            class_assigned=school_class,
                            session=session,
                            term=term,
                            ca1_marks=ca1_marks,
                            ca2_marks=ca2_marks,
                            home_work_marks=home_work_marks,
                            activity_marks=activity_marks,
                            exam_marks=exam_marks,
                        )

                messages.success(request, "Results processed. Check warnings for existing records.")
                return redirect('result_entry')

            return render(request, 'src/result_entry.html', {
                'subjects': subjects,
                'classes': classes,
                'sessions': sessions,
                'terms': terms,
                'students': students,
                'selected_term': term,
                'selected_session': session,
                'selected_class': school_class,
                'selected_subject': subject,
            })

    return render(request, 'src/result_entry.html', {
        'subjects': subjects,
        'classes': classes,
        'sessions': sessions,
        'terms': terms,
    })


@login_required(login_url='login')
def result_update(request):
    subjects = Subject.objects.all()
    classes = SchoolClass.objects.all()
    sessions = Session.objects.all()
    terms = Term.objects.all()

    if request.method == 'POST':
        term_id = request.POST.get('term')
        session_id = request.POST.get('session')
        class_id = request.POST.get('class_assigned')
        subject_id = request.POST.get('subject')

        if term_id and session_id and class_id and subject_id:
            term = get_object_or_404(Term, pk=term_id)
            session = get_object_or_404(Session, pk=session_id)
            school_class = get_object_or_404(SchoolClass, pk=class_id)
            subject = get_object_or_404(Subject, pk=subject_id)
            students = Student.objects.filter(enrolled_class=school_class)

            # Load existing results for the selected session, term, class, and subject
            results = Result.objects.filter(
                session=session, term=term, class_assigned=school_class, subject=subject
            )

            # If the save button is clicked, update the results
            if 'save_results' in request.POST:
                for result in results:
                    # Ensure that marks are integers, defaulting to 0 if invalid or missing
                    ca1_marks = request.POST.get(f'ca1_{result.student.id}', '0')
                    ca2_marks = request.POST.get(f'ca2_{result.student.id}', '0')
                    home_work_marks = request.POST.get(f'home_work_{result.student.id}', '0')
                    activity_marks = request.POST.get(f'activity_{result.student.id}', '0')
                    exam_marks = request.POST.get(f'exam_{result.student.id}', '0')

                    # Use int() conversion directly and handle exceptions
                    try:
                        result.ca1_marks = int(ca1_marks)
                    except ValueError:
                        result.ca1_marks = 0

                    try:
                        result.ca2_marks = int(ca2_marks)
                    except ValueError:
                        result.ca2_marks = 0

                    try:
                        result.home_work_marks = int(home_work_marks)
                    except ValueError:
                        result.home_work_marks = 0

                    try:
                        result.activity_marks = int(activity_marks)
                    except ValueError:
                        result.activity_marks = 0

                    try:
                        result.exam_marks = int(exam_marks)
                    except ValueError:
                        result.exam_marks = 0

                    result.save()

                messages.success(request, "Results successfully updated.")
                return redirect('result_update')

            return render(request, 'src/result_update.html', {
                'subjects': subjects,
                'classes': classes,
                'sessions': sessions,
                'terms': terms,
                'students': students,
                'results': results,
                'selected_term': term,
                'selected_session': session,
                'selected_class': school_class,
                'selected_subject': subject,
            })

    return render(request, 'src/result_update.html', {
        'subjects': subjects,
        'classes': classes,
        'sessions': sessions,
        'terms': terms,
    })



@login_required(login_url='login')
def download_template(request):
    classes = SchoolClass.objects.all()

    if request.method == 'POST':
        class_id = request.POST.get('class_assigned')
        school_class = get_object_or_404(SchoolClass, pk=class_id)
        students = Student.objects.filter(enrolled_class=school_class)

        # Create an in-memory output file for the new workbook.
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet()

        # Define the headers
        worksheet.write('A1', 'Student ID')
        worksheet.write('B1', 'Student Name')
        worksheet.write('C1', '1st CA Marks')
        worksheet.write('D1', '2nd CA Marks')
        worksheet.write('E1', 'Home Work Marks')
        worksheet.write('F1', 'Activity Marks')
        worksheet.write('G1', 'Exam Marks')
        worksheet.write('H1', 'Admission Number')

        # Write student data
        for row_num, student in enumerate(students, start=2):
            worksheet.write(f'A{row_num}', student.id)
            worksheet.write(f'B{row_num}', f"{student.first_name} {student.last_name}")
            worksheet.write(f'C{row_num}', 0)  # Placeholder for 1st CA Marks
            worksheet.write(f'D{row_num}', 0)  # Placeholder for 2nd CA Marks
            worksheet.write(f'E{row_num}', 0)  # Placeholder for Home Work Marks
            worksheet.write(f'F{row_num}', 0)  # Placeholder for Activity Marks
            worksheet.write(f'G{row_num}', 0)  # Placeholder for Exam Marks
            worksheet.write(f'H{row_num}', student.admission_number)  # Placeholder for Exam Marks

        workbook.close()
        output.seek(0)

        # Return the Excel file as a response
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=template_{school_class.name}_{school_class.arm}.xlsx'

        return response

    return render(request, 'src/download_template.html', {'classes': classes})



@login_required(login_url='login')
def upload_results(request):
    subjects = Subject.objects.all()
    classes = SchoolClass.objects.all()
    sessions = Session.objects.all()
    terms = Term.objects.all()

    if request.method == 'POST':
        session_id = request.POST.get('session')
        term_id = request.POST.get('term')
        class_id = request.POST.get('class_assigned')
        subject_id = request.POST.get('subject')
        excel_file = request.FILES.get('excel_file')

        if session_id and term_id and class_id and subject_id and excel_file:
            session = get_object_or_404(Session, pk=session_id)
            term = get_object_or_404(Term, pk=term_id)
            school_class = get_object_or_404(SchoolClass, pk=class_id)
            subject = get_object_or_404(Subject, pk=subject_id)

            workbook = openpyxl.load_workbook(excel_file)
            worksheet = workbook.active

            for row in worksheet.iter_rows(min_row=2, values_only=True):
                student_id, ca1_marks, ca2_marks, home_work_marks, activity_marks, exam_marks = row[0], row[2], row[3], row[4], row[5], row[6]

                try:
                    student = Student.objects.get(id=student_id, enrolled_class=school_class)

                    # Check if result already exists
                    if not Result.objects.filter(
                        student=student,
                        subject=subject,
                        class_assigned=school_class,
                        session=session,
                        term=term
                    ).exists():
                        # Create a new result
                        Result.objects.create(
                            student=student,
                            subject=subject,
                            class_assigned=school_class,
                            session=session,
                            term=term,
                            ca1_marks=int(ca1_marks),
                            ca2_marks=int(ca2_marks),
                            home_work_marks=int(home_work_marks),
                            activity_marks=int(activity_marks),
                            exam_marks=int(exam_marks),
                        )
                    else:
                        messages.warning(request, f"Result for {student.first_name} {student.last_name} already exists.")
                except Student.DoesNotExist:
                    messages.error(request, f"Student with ID {student_id} not found.")

            messages.success(request, "Results successfully uploaded.")
            return redirect('upload_results')

    return render(request, 'src/upload_results.html', {
        'subjects': subjects,
        'classes': classes,
        'sessions': sessions,
        'terms': terms,
    })



@login_required(login_url='login')
def upload_missed_results(request):
    subjects = Subject.objects.all()
    classes = SchoolClass.objects.all()
    sessions = Session.objects.all()
    terms = Term.objects.all()

    if request.method == 'POST':
        session_id = request.POST.get('session')
        term_id = request.POST.get('term')
        class_id = request.POST.get('class_assigned')
        subject_id = request.POST.get('subject')
        excel_file = request.FILES.get('excel_file')

        if session_id and term_id and class_id and subject_id and excel_file:
            session = get_object_or_404(Session, pk=session_id)
            term = get_object_or_404(Term, pk=term_id)
            school_class = get_object_or_404(SchoolClass, pk=class_id)
            subject = get_object_or_404(Subject, pk=subject_id)

            workbook = openpyxl.load_workbook(excel_file)
            worksheet = workbook.active

            for row in worksheet.iter_rows(min_row=2, values_only=True):
                student_id, ca1_marks, ca2_marks, home_work_marks, activity_marks, exam_marks = row[0], row[2], row[3], row[4], row[5], row[6]

                try:
                    student = Student.objects.get(id=student_id, enrolled_class=school_class)

                    # Check if result does not already exist (i.e., missed student)
                    if not Result.objects.filter(
                        student=student,
                        subject=subject,
                        class_assigned=school_class,
                        session=session,
                        term=term
                    ).exists():
                        # Create a new result for the missed student
                        Result.objects.create(
                            student=student,
                            subject=subject,
                            class_assigned=school_class,
                            session=session,
                            term=term,
                            ca1_marks=int(ca1_marks),
                            ca2_marks=int(ca2_marks),
                            home_work_marks=int(home_work_marks),
                            activity_marks=int(activity_marks),
                            exam_marks=int(exam_marks),
                        )
                        messages.success(request, f"Result for {student.first_name} {student.last_name} added successfully.")
                    else:
                        messages.warning(request, f"Result for {student.first_name} {student.last_name} already exists.")

                except Student.DoesNotExist:
                    messages.error(request, f"Student with ID {student_id} not found in the selected class.")

            return redirect('upload_missed_results')

    return render(request, 'src/upload_missed_results.html', {
        'subjects': subjects,
        'classes': classes,
        'sessions': sessions,
        'terms': terms,
    })


@login_required(login_url='login')
def download_behavioral_template(request):
    classes = SchoolClass.objects.all()

    if request.method == 'POST':
        class_id = request.POST.get('class_assigned')
        school_class = get_object_or_404(SchoolClass, pk=class_id)
        students = Student.objects.filter(enrolled_class=school_class)

        # Create an in-memory output file for the new workbook.
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet()

        # Define the headers
        worksheet.write('A1', 'Student ID')
        worksheet.write('B1', 'Student Name')
        worksheet.write('C1', 'Conduct')
        worksheet.write('D1', 'Punctuality')
        worksheet.write('E1', 'Dedication')
        worksheet.write('F1', 'Participation')
        worksheet.write('G1', 'Hospitality')
        worksheet.write('H1', 'Neatness')
        worksheet.write('I1', 'Creativity')
        worksheet.write('J1', 'Physical')

        # Write student data
        for row_num, student in enumerate(students, start=2):
            worksheet.write(f'A{row_num}', student.id)
            worksheet.write(f'B{row_num}', f"{student.first_name} {student.last_name}")
            worksheet.write(f'C{row_num}', '')  # Placeholder for Conduct
            worksheet.write(f'D{row_num}', '')  # Placeholder for Punctuality
            worksheet.write(f'E{row_num}', '')  # Placeholder for Dedication
            worksheet.write(f'F{row_num}', '')  # Placeholder for Participation
            worksheet.write(f'G{row_num}', '')  # Placeholder for Hospitality
            worksheet.write(f'H{row_num}', '')  # Placeholder for Neatness
            worksheet.write(f'I{row_num}', '')  # Placeholder for Creativity
            worksheet.write(f'J{row_num}', '')  # Placeholder for Physical

        workbook.close()
        output.seek(0)

        # Return the Excel file as a response
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=behavioral_template_{school_class.name}.xlsx'

        return response

    return render(request, 'src/download_behavioral_template.html', {'classes': classes})




@login_required(login_url='login')
def upload_behavioral_assessments(request):
    classes = SchoolClass.objects.all()
    sessions = Session.objects.all()
    terms = Term.objects.all()

    if request.method == 'POST':
        session_id = request.POST.get('session')
        term_id = request.POST.get('term')
        class_id = request.POST.get('class_assigned')
        excel_file = request.FILES.get('excel_file')

        if session_id and term_id and class_id and excel_file:
            session = get_object_or_404(Session, pk=session_id)
            term = get_object_or_404(Term, pk=term_id)
            school_class = get_object_or_404(SchoolClass, pk=class_id)

            workbook = openpyxl.load_workbook(excel_file)
            worksheet = workbook.active

            for row in worksheet.iter_rows(min_row=2, values_only=True):
                # Ensure you are reading the correct columns
                student_id = row[0]
                conduct = row[2]
                punctuality = row[3]
                dedication = row[4]
                participation = row[5]
                hospitality = row[6]
                neatness = row[7]
                creativity = row[8]
                physical = row[9]

                try:
                    student = Student.objects.get(id=student_id, enrolled_class=school_class)

                    # Check if behavioral assessment already exists
                    behavioral_assessment, created = StudentBehaviouralAssessment.objects.update_or_create(
                        student=student,
                        session=session,
                        term=term,
                        school_class=school_class,
                        defaults={
                            'conduct': int(conduct) if conduct and str(conduct).isdigit() else 0,
                            'punctuality': int(punctuality) if punctuality and str(punctuality).isdigit() else 0,
                            'dedication': int(dedication) if dedication and str(dedication).isdigit() else 0,
                            'participation': int(participation) if participation and str(participation).isdigit() else 0,
                            'hospitality': int(hospitality) if hospitality and str(hospitality).isdigit() else 0,
                            'neatness': int(neatness) if neatness and str(neatness).isdigit() else 0,
                            'creativity': int(creativity) if creativity and str(creativity).isdigit() else 0,
                            'physical': int(physical) if physical and str(physical).isdigit() else 0,
                        }
                    )

                    if created:
                        messages.success(request, f"Assessment for {student.first_name} {student.last_name} added successfully.")
                    else:
                        messages.info(request, f"Assessment for {student.first_name} {student.last_name} updated successfully.")

                except Student.DoesNotExist:
                    messages.error(request, f"Student with ID {student_id} not found in the selected class.")
                except ValueError as e:
                    messages.error(request, f"Error processing row for student ID {student_id}: {e}")

            return redirect('upload_behavioral_assessments')

    return render(request, 'src/upload_behavioral_assessments.html', {
        'classes': classes,
        'sessions': sessions,
        'terms': terms,
    })




@login_required(login_url='login')
def view_behavioral_assessments(request):
    classes = SchoolClass.objects.all()
    sessions = Session.objects.all()
    terms = Term.objects.all()
    assessments = None

    if request.method == 'POST':
        session_id = request.POST.get('session')
        term_id = request.POST.get('term')
        class_id = request.POST.get('class_assigned')

        if session_id and term_id and class_id:
            session = get_object_or_404(Session, pk=session_id)
            term = get_object_or_404(Term, pk=term_id)
            school_class = get_object_or_404(SchoolClass, pk=class_id)

            # Query for behavioral assessments based on selected filters
            assessments = StudentBehaviouralAssessment.objects.filter(
                session=session,
                term=term,
                school_class=school_class
            )

            if not assessments.exists():
                messages.warning(request, "No behavioral assessments found for the selected criteria.")

    return render(request, 'src/view_behavioral_assessments.html', {
        'classes': classes,
        'sessions': sessions,
        'terms': terms,
        'assessments': assessments,
    })







@login_required(login_url='login')
def select_class_for_result(request):
    classes = SchoolClass.objects.all()
    sessions = Session.objects.all()
    terms = Term.objects.all()

    if request.method == 'POST':
        session_id = request.POST.get('session')
        term_id = request.POST.get('term')
        class_id = request.POST.get('class_assigned')

        if session_id and term_id and class_id:
            return redirect('display_class_results', session_id=session_id, term_id=term_id, class_id=class_id)

    return render(request, 'src/select_class_for_result.html', {
        'classes': classes,
        'sessions': sessions,
        'terms': terms,
    })



@login_required(login_url='login')
def display_class_results(request, session_id, term_id, class_id):
    session = get_object_or_404(Session, pk=session_id)
    term = get_object_or_404(Term, pk=term_id)
    school_class = get_object_or_404(SchoolClass, pk=class_id)
    students = Student.objects.filter(enrolled_class=school_class)
    school_config = SchoolConfig.objects.last()

    results_data = []

    for student in students:
        # Get all results for the student
        results = Result.objects.filter(
            session=session,
            term=term,
            class_assigned=school_class,
            student=student
        )

        # Compute total score and average for the student
        total_score = sum(result.ca1_marks + result.ca2_marks + result.home_work_marks + result.activity_marks + result.exam_marks for result in results)
        num_subjects = results.count()
        average_score = total_score / num_subjects if num_subjects > 0 else 0
        
        
        if 76 <= average_score <= 100:
            overall_grade = "A+"
        elif 70 <= average_score < 76:  # Use `< 76` to include fractional values
            overall_grade = "A"
        elif 65 <= average_score < 70:  # Use `< 70` and so on
            overall_grade = "A-"
        elif 60 <= average_score < 65:
            overall_grade = "B+"
        elif 55 <= average_score < 60:
            overall_grade = "B"
        elif 50 <= average_score < 55:
            overall_grade = "B-"
        elif 46 <= average_score < 50:
            overall_grade = "C+"
        elif 43 <= average_score < 46:
            overall_grade = "C"
        elif 39 <= average_score < 43:
            overall_grade = "C-"
        elif 0 <= average_score < 39:
            overall_grade = "F"
        else:
            overall_grade = "Invalid score"



        # Fetch behavioral assessment
        behavioral_assessment = StudentBehaviouralAssessment.objects.filter(
            session=session,
            term=term,
            school_class=school_class,
            student=student
        ).first()

        # Retrieve the fee structure for the student's class, session, and term
        fee_structure = FeeStructure.objects.filter(
            class_assigned=school_class,
            session=session,
            term=term
        ).first()

        # Initialize fee details
        total_fee = fee_structure.amount if fee_structure else 0

        # Retrieve payment made by the student for the same session and term
        payment = Payment.objects.filter(
            student=student,
            session=session,
            term=term
        ).aggregate(amount_paid=models.Sum('amount_paid'))['amount_paid'] or 0

        # Calculate outstanding balance
        outstanding_balance = total_fee - payment

        # Comments based on average score
        if 76 <= average_score <= 100:
            english_comment = "AN EXCELLENT PERFORMANCE, KEEP IT UP."
            arabic_comment_male = "فاز بتقدير ممتاز ويرجى له التفوق في الفترات القادمة"
            arabic_comment_female = "فازت بتقدير ممتاز ويرجى لها التفوق في الفترات القادمة"
        elif 70 <= average_score <= 76:
            english_comment = "AN EXCELLENT PERFORMANCE, KEEP IT UP."
            arabic_comment_male = "فاز بتقدير ممتاز ويرجى له التفوق في الفترات القادمة"
            arabic_comment_female = "فازت بتقدير ممتاز ويرجى لها التفوق في الفترات القادمة"
        elif 65 <= average_score <= 70:
            english_comment = "AN EXCELLENT PERFORMANCE, KEEP IT UP."
            arabic_comment_male = "فاز بتقدير ممتاز ويرجى له التفوق في الفترات القادمة"
            arabic_comment_female = "فازت بتقدير ممتاز ويرجى لها التفوق في الفترات القادمة"
        elif 60 <= average_score <= 65:
            english_comment = "A VERY GOOD RESULT, PUT IN MORE EFFORT."
            arabic_comment_male = "فاز بتقدير جيد جدا ويرجى له التقدم في الفترة المقبلة"
            arabic_comment_female = "فازت بتقدير جيد جدا ويرجى لها التقدم في الفترة المقبلة"
        elif 55 <= average_score <= 60:
            english_comment = "A VERY GOOD RESULT, PUT IN MORE EFFORT."
            arabic_comment_male = "فاز بتقدير جيد جدا ويرجى له التقدم في الفترة المقبلة"
            arabic_comment_female = "فازت بتقدير جيد جدا ويرجى لها التقدم في الفترة المقبلة"
        elif 50 <= average_score <= 55:
            english_comment = "A VERY GOOD RESULT, PUT IN MORE EFFORT."
            arabic_comment_male = "فاز بتقدير جيد جدا ويرجى له التقدم في الفترة المقبلة"
            arabic_comment_female = "فازت بتقدير جيد جدا ويرجى لها التقدم في الفترة المقبلة"
        elif 46 <= average_score <= 50:
            english_comment = "A GOOD RESULT, TRY HARDER NEXT TERM."
            arabic_comment_male = "فاز بتقدير جيد ويرجى له الجهد الكبير في الفترة المقبلة"
            arabic_comment_female = "فازت بتقدير جيد ويرجى لها الجهد الكبير في الفترة المقبلة"
        elif 43 <= average_score <= 46:
            english_comment = "A GOOD RESULT, TRY HARDER NEXT TERM."
            arabic_comment_male = "فاز بتقدير جيد ويرجى له الجهد الكبير في الفترة المقبلة"
            arabic_comment_female = "فازت بتقدير جيد ويرجى لها الجهد الكبير في الفترة المقبلة"
        elif 39 <= average_score <= 43:
            english_comment = "A GOOD RESULT, TRY HARDER NEXT TERM."
            arabic_comment_male = "فاز بتقدير جيد ويرجى له الجهد الكبير في الفترة المقبلة"
            arabic_comment_female = "فازت بتقدير جيد ويرجى لها الجهد الكبير في الفترة المقبلة"
        elif 0 <= average_score <= 39:
            english_comment = "A SATISFACTORY RESULT, TRY TO IMPROVE NEXT TERM."
            arabic_comment_male = "تقدير ضعيف،يرجى منه التقدم"
            arabic_comment_female = "تقدير ضعيف، يرجى منها التقدم"
        else:
            english_comment = "Invalid score."
            arabic_comment_male = arabic_comment_female = "درجة غير صالحة."

        # Determine gender-specific Arabic comment
        if student.gender == "Male":
            comments = f"{english_comment}\n{arabic_comment_male}"
        elif student.gender == "Female":
            comments = f"{english_comment}\n{arabic_comment_female}"
        else:
            comments = f"{english_comment}\nUnknown gender for Arabic comment."


        # Prepare results for the student and calculate positions for each subject
        student_results = []
        for result in results:
            # Fetch all results for the same subject, term, session, and class to calculate position
            subject_results = Result.objects.filter(
                subject=result.subject,
                session=session,
                term=term,
                class_assigned=school_class
            )

      
            # Convert position to ordinal representation
            position_ordinal = result.subject_position

            # Append each subject's result and position to the student's result data
            student_results.append({
                'subject': result.subject,
                'ca1_marks': result.ca1_marks,
                'ca2_marks': result.ca2_marks,
                'home_work_marks': result.home_work_marks,
                'activity_marks': result.activity_marks,
                'exam_marks': result.exam_marks,
                'total_marks': result.ca1_marks + result.ca2_marks + result.home_work_marks + result.activity_marks + result.exam_marks,
                'grade': result.grade,
                'position': position_ordinal  # Use the ordinal position
            })

        results_data.append({
            'student': student,
            'results': student_results,
            'total_score': total_score,
            'average_score': average_score,
            'overall_grade': overall_grade,
            'behavioral_assessment': behavioral_assessment,
            'comments': comments,
            'total_fee': total_fee,
            'amount_paid': payment,
            'outstanding_balance': outstanding_balance
        })

    return render(request, 'src/display_class_results.html', {
        'session': session,
        'term': term,
        'school_class': school_class,
        'results_data': results_data,
        'school_config': school_config,
    })




@login_required(login_url='login')
def student_result_search(request):
    sessions = Session.objects.all()
    terms = Term.objects.all()
    classes = SchoolClass.objects.all()
    school_config = SchoolConfig.objects.last()  # Retrieve the latest school config

    if request.method == 'GET':
        return render(request, 'src/student_result_search.html', {
            'sessions': sessions,
            'terms': terms,
            'classes': classes,
            'school_config': school_config,  # Pass the school config to the template
        })



@login_required(login_url='login')
def view_student_result(request):
    # Get data from form (session, term, class, and admission number)
    session_id = request.GET.get('session_id')
    term_id = request.GET.get('term_id')
    class_id = request.GET.get('class_id')
    admission_number = request.GET.get('admission_number')

    # Fetch session, term, class, and student
    session = get_object_or_404(Session, pk=session_id)
    term = get_object_or_404(Term, pk=term_id)
    school_class = get_object_or_404(SchoolClass, pk=class_id)
    student = get_object_or_404(Student, admission_number=admission_number)

    # Fetch results for the student in the selected session, term, and class
    results = Result.objects.filter(
        session=session,
        term=term,
        class_assigned=school_class,
        student=student
    )

    # Calculate total score and average
    total_score = sum(result.ca1_marks + result.ca2_marks + result.home_work_marks + result.activity_marks + result.exam_marks for result in results)
    num_subjects = results.count()
    average_score = total_score / num_subjects if num_subjects > 0 else 0
    

    # Calculate subject position for each result
    student_results = []
    for result in results:
        # Get all results for the same subject, class, term, and session
        subject_results = Result.objects.filter(
            subject=result.subject,
            session=session,
            term=term,
            class_assigned=school_class
        )

        # Sort results by total marks (CA1 + CA2 + Exam)
        subject_results = sorted(subject_results, key=lambda r: r.ca1_marks + r.ca2_marks + r.home_work_marks + r.activity_marks + r.exam_marks, reverse=True)

        # Find the position of the current student
        position = subject_results.index(result) + 1
        position_with_ordinal = ordinal(position)  # Convert to ordinal

        # Append each subject's result and position to the student's result data
        student_results.append({
            'subject': result.subject,
            'ca1_marks': result.ca1_marks,
            'ca2_marks': result.ca2_marks,
            'home_work_marks': result.home_work_marks,
            'activity_marks': result.activity_marks,
            'exam_marks': result.exam_marks,
            'total_marks': result.ca1_marks + result.ca2_marks + result.exam_marks,
            'grade': result.grade,
            'position': position_with_ordinal  # Ordinal position
        })

    # Fetch behavioral assessment
    behavioral_assessment = StudentBehaviouralAssessment.objects.filter(
        session=session,
        term=term,
        school_class=school_class,
        student=student
    ).first()

    # Retrieve the fee structure for the student's class, session, and term
    fee_structure = FeeStructure.objects.filter(
        class_assigned=school_class,
        session=session,
        term=term
    ).first()

    # Total fee from the structure (if exists)
    total_fee = fee_structure.amount if fee_structure else 0

    # Retrieve payment made by the student for the same session and term
    payment = Payment.objects.filter(
        student=student,
        session=session,
        term=term
    ).aggregate(amount_paid=Sum('amount_paid'))['amount_paid'] or 0

    # Calculate outstanding balance
    outstanding_balance = total_fee - payment

    # Comments based on average score
    if 76 <= average_score <= 100:
        comments = "AN EXCELLENT PERFORMANCE, KEEP IT UP."
    elif 70 <= average_score <= 75:
        comments = "AN EXCELLENT PERFORMANCE, KEEP IT UP."
    elif 65 <= average_score <= 69:
        comments = "AN EXCELLENT PERFORMANCE, KEEP IT UP."
    elif 60 <= average_score <= 64:
        comments = "A VERY GOOD RESULT, PUT IN MORE EFFORT."
    elif 55 <= average_score <= 59:
        comments = "A VERY GOOD RESULT, PUT IN MORE EFFORT."
    elif 50 <= average_score <= 54:
        comments = "A VERY GOOD RESULT, PUT IN MORE EFFORT."
    elif 46 <= average_score <= 49:
        comments = "A GOOD RESULT, TRY HARDER NEXT TERM."
    elif 43 <= average_score <= 45:
        comments = "A GOOD RESULT, TRY HARDER NEXT TERM."
    elif 39 <= average_score <= 42:
        comments = "A GOOD RESULT, TRY HARDER NEXT TERM."
    elif 0 <= average_score <= 38:
        comments = "A SATISFACTORY RESULT, TRY TO IMPROVE NEXT TERM."
    else:
        comments = "Invalid score."

    # Prepare data for the template
    school_config = SchoolConfig.objects.last()

    context = {
        'student': student,
        'results': student_results,  # Using updated results with positions
        'total_score': total_score,
        'average_score': average_score,
        'behavioral_assessment': behavioral_assessment,
        'total_fee': total_fee,
        'amount_paid': payment,
        'outstanding_balance': outstanding_balance,
        'school_config': school_config,  # For header and signature images
        'comments': comments,
    }

    return render(request, 'src/view_student_result.html', context)


def result_checker(request):

    if request.method == 'POST':
        # Get form data
        session_id = request.POST.get('session')
        term_id = request.POST.get('term')
        admission_number = request.POST.get('admission_number')
        token_code = request.POST.get('token')

        student = Student.objects.get(admission_number=admission_number)

        session = Session.objects.get(id=session_id)
        term = Term.objects.get(id=term_id)
        try:
            token = Token.objects.get(token_code=token_code)
            token.use_token(student, session, term)
            print("Token successfully used!")
            return redirect('display_single_result', session_id=session_id, term_id=term_id, student_id=student.id, token_code=token_code)

        except Token.DoesNotExist:
        # Handle case where the token is invalid
            return render(request, 'src/error.html', {
                'message': 'Invalid token or usage count exceeded.',
                
            })
        except ValueError as e:
                    # Handle case where the token is invalid
            return render(request, 'src/error.html', {
                'message': f"Error: {str(e)}",

            })

    context = {
        'sessions': Session.objects.all(),
        'terms': Term.objects.all(),
    }

    return render(request, 'src/result_checker.html', context)



def display_single_result(request, session_id, term_id, student_id, token_code):
    session = get_object_or_404(Session, pk=session_id)
    term = get_object_or_404(Term, pk=term_id)
    student = get_object_or_404(Student, id=student_id)
    school_class = get_object_or_404(SchoolClass, id=student.enrolled_class_id)
    token = get_object_or_404(Token, token_code=token_code)
    school_config = SchoolConfig.objects.last()

    results_data = []
      
    if token.session == session and token.term == term and token.associated_student == student and token.usage_count < token.max_usage:
        # Get all results for the student
        results = Result.objects.filter(session=session, term=term, student=student, class_assigned=school_class)

        # Compute total score and average for the student
        total_score = sum(
            result.ca1_marks + result.ca2_marks + result.home_work_marks +
            result.activity_marks + result.exam_marks for result in results
        )
        num_subjects = results.count()
        average_score = total_score / num_subjects if num_subjects > 0 else 0
        

        # Determine the overall grade
        if 76 <= average_score <= 100:
            overall_grade = "A+"
        elif 70 <= average_score < 76:  # Use `< 76` to include fractional values
            overall_grade = "A"
        elif 65 <= average_score < 70:  # Use `< 70` and so on
            overall_grade = "A-"
        elif 60 <= average_score < 65:
            overall_grade = "B+"
        elif 55 <= average_score < 60:
            overall_grade = "B"
        elif 50 <= average_score < 55:
            overall_grade = "B-"
        elif 46 <= average_score < 50:
            overall_grade = "C+"
        elif 43 <= average_score < 46:
            overall_grade = "C"
        elif 39 <= average_score < 43:
            overall_grade = "C-"
        elif 0 <= average_score < 39:
            overall_grade = "F"
        else:
            overall_grade = "Invalid score"


        # Fetch behavioral assessment
        behavioral_assessment = StudentBehaviouralAssessment.objects.filter(
            session=session, term=term, school_class=school_class, student=student
        ).first()

        # Retrieve fee structure for the student's class, session, and term
        fee_structure = FeeStructure.objects.filter(
            class_assigned=school_class, session=session, term=term
        ).first()

        total_fee = fee_structure.amount if fee_structure else 0

        # Retrieve payment made by the student for the same session and term
        payment = Payment.objects.filter(
            student=student, session=session, term=term
        ).aggregate(amount_paid=Sum('amount_paid'))['amount_paid'] or 0

        # Calculate outstanding balance
        outstanding_balance = total_fee - payment

        # Comments based on average score
        if 76 <= average_score <= 100:
            english_comment = "AN EXCELLENT PERFORMANCE, KEEP IT UP."
            arabic_comment_male = "فاز بتقدير ممتاز ويرجى له التفوق في الفترات القادمة"
            arabic_comment_female = "فازت بتقدير ممتاز ويرجى لها التفوق في الفترات القادمة"
        elif 70 <= average_score <= 76:
            english_comment = "AN EXCELLENT PERFORMANCE, KEEP IT UP."
            arabic_comment_male = "فاز بتقدير ممتاز ويرجى له التفوق في الفترات القادمة"
            arabic_comment_female = "فازت بتقدير ممتاز ويرجى لها التفوق في الفترات القادمة"
        elif 65 <= average_score <= 70:
            english_comment = "AN EXCELLENT PERFORMANCE, KEEP IT UP."
            arabic_comment_male = "فاز بتقدير ممتاز ويرجى له التفوق في الفترات القادمة"
            arabic_comment_female = "فازت بتقدير ممتاز ويرجى لها التفوق في الفترات القادمة"
        elif 60 <= average_score <= 65:
            english_comment = "A VERY GOOD RESULT, PUT IN MORE EFFORT."
            arabic_comment_male = "فاز بتقدير جيد جدا ويرجى له التقدم في الفترة المقبلة"
            arabic_comment_female = "فازت بتقدير جيد جدا ويرجى لها التقدم في الفترة المقبلة"
        elif 55 <= average_score <= 60:
            english_comment = "A VERY GOOD RESULT, PUT IN MORE EFFORT."
            arabic_comment_male = "فاز بتقدير جيد جدا ويرجى له التقدم في الفترة المقبلة"
            arabic_comment_female = "فازت بتقدير جيد جدا ويرجى لها التقدم في الفترة المقبلة"
        elif 50 <= average_score <= 55:
            english_comment = "A VERY GOOD RESULT, PUT IN MORE EFFORT."
            arabic_comment_male = "فاز بتقدير جيد جدا ويرجى له التقدم في الفترة المقبلة"
            arabic_comment_female = "فازت بتقدير جيد جدا ويرجى لها التقدم في الفترة المقبلة"
        elif 46 <= average_score <= 50:
            english_comment = "A GOOD RESULT, TRY HARDER NEXT TERM."
            arabic_comment_male = "فاز بتقدير جيد ويرجى له الجهد الكبير في الفترة المقبلة"
            arabic_comment_female = "فازت بتقدير جيد ويرجى لها الجهد الكبير في الفترة المقبلة"
        elif 43 <= average_score <= 46:
            english_comment = "A GOOD RESULT, TRY HARDER NEXT TERM."
            arabic_comment_male = "فاز بتقدير جيد ويرجى له الجهد الكبير في الفترة المقبلة"
            arabic_comment_female = "فازت بتقدير جيد ويرجى لها الجهد الكبير في الفترة المقبلة"
        elif 39 <= average_score <= 43:
            english_comment = "A GOOD RESULT, TRY HARDER NEXT TERM."
            arabic_comment_male = "فاز بتقدير جيد ويرجى له الجهد الكبير في الفترة المقبلة"
            arabic_comment_female = "فازت بتقدير جيد ويرجى لها الجهد الكبير في الفترة المقبلة"
        elif 0 <= average_score <= 39:
            english_comment = "A SATISFACTORY RESULT, TRY TO IMPROVE NEXT TERM."
            arabic_comment_male = "تقدير ضعيف،يرجى منه التقدم"
            arabic_comment_female = "تقدير ضعيف، يرجى منها التقدم"
        else:
            english_comment = "Invalid score."
            arabic_comment_male = arabic_comment_female = "درجة غير صالحة."

        # Determine gender-specific Arabic comment
        if student.gender == "Male":
            comments = f"{english_comment}\n{arabic_comment_male}"
        elif student.gender == "Female":
            comments = f"{english_comment}\n{arabic_comment_female}"
        else:
            comments = f"{english_comment}\nUnknown gender for Arabic comment."

        # Prepare results for the student and calculate positions for each subject
        student_results = []
        for result in results:
            # Fetch all results for the same subject, term, session, and class to calculate position
            subject_results = Result.objects.filter(
                subject=result.subject,
                session=session,
                term=term,
                class_assigned=school_class
            )

       
            # Convert position to ordinal representation
            position_ordinal = result.subject_position

            # Append each subject's result and position to the student's result data
            student_results.append({
                'subject': result.subject,
                'ca1_marks': result.ca1_marks,
                'ca2_marks': result.ca2_marks,
                'home_work_marks': result.home_work_marks,
                'activity_marks': result.activity_marks,
                'exam_marks': result.exam_marks,
                'total_marks': result.ca1_marks + result.ca2_marks + result.home_work_marks + result.activity_marks + result.exam_marks,
                'grade': result.grade,
                'position': position_ordinal  # Use the ordinal position
            })

        results_data.append({
            'student': student,
            'results': student_results,
            'total_score': total_score,
            'average_score': average_score,
            'overall_grade': overall_grade,
            'behavioral_assessment': behavioral_assessment,
            'comments': comments,
            'total_fee': total_fee,
            'amount_paid': payment,
            'outstanding_balance': outstanding_balance
        })

        return render(request, 'src/display_single_result.html', {
            'session': session,
            'term': term,
            'school_class': school_class,
            'results_data': results_data,
            'school_config': school_config,
        })

    else:
        # Handle case where the token is invalid
        return render(request, 'src/error.html', {
            'message': 'Invalid token or usage count exceeded.',
            'school_config': school_config,
        })


@login_required(login_url='login')
def select_class_for_result_summary(request):
    classes = SchoolClass.objects.all()
    sessions = Session.objects.all()
    terms = Term.objects.all()

    if request.method == 'POST':
        session_id = request.POST.get('session')
        term_id = request.POST.get('term')
        class_id = request.POST.get('class_assigned')

        if session_id and term_id and class_id:
            return redirect('display_class_results_summary', session_id=session_id, term_id=term_id, class_id=class_id)

    return render(request, 'src/select_class_for_result.html', {
        'classes': classes,
        'sessions': sessions,
        'terms': terms,
    })



@login_required(login_url='login')
def display_class_results_summary(request, session_id, term_id, class_id):
    session = get_object_or_404(Session, pk=session_id)
    term = get_object_or_404(Term, pk=term_id)
    school_class = get_object_or_404(SchoolClass, pk=class_id)
    students = Student.objects.filter(enrolled_class=school_class)
    school_config = SchoolConfig.objects.last()

    results_data = []

    for student in students:
        # Get all results for the student
        results = Result.objects.filter(
            session=session,
            term=term,
            class_assigned=school_class,
            student=student
        )

        # Compute total score and average for the student
        total_score = sum(result.ca1_marks + result.ca2_marks + result.home_work_marks + result.activity_marks + result.exam_marks for result in results)
        num_subjects = results.count()
        average_score = total_score / num_subjects if num_subjects > 0 else 0
        
        if 76 <= average_score <= 100:
            overall_grade = "A+"
        elif 70 <= average_score < 76:  # Use `< 76` to include fractional values
            overall_grade = "A"
        elif 65 <= average_score < 70:  # Use `< 70` and so on
            overall_grade = "A-"
        elif 60 <= average_score < 65:
            overall_grade = "B+"
        elif 55 <= average_score < 60:
            overall_grade = "B"
        elif 50 <= average_score < 55:
            overall_grade = "B-"
        elif 46 <= average_score < 50:
            overall_grade = "C+"
        elif 43 <= average_score < 46:
            overall_grade = "C"
        elif 39 <= average_score < 43:
            overall_grade = "C-"
        elif 0 <= average_score < 39:
            overall_grade = "F"
        else:
            overall_grade = "Invalid score"


       
        # Prepare results for the student and calculate positions for each subject
        student_results = []
        for result in results:
            # Fetch all results for the same subject, term, session, and class to calculate position
            subject_results = Result.objects.filter(
                subject=result.subject,
                session=session,
                term=term,
                class_assigned=school_class
            )

            # Calculate the total marks for each student and sort them
            subject_results = sorted(subject_results, key=lambda r: r.ca1_marks + r.ca2_marks + r.exam_marks, reverse=True)



            # Append each subject's result and position to the student's result data
            student_results.append({
                'subject': result.subject,
                'ca1_marks': result.ca1_marks,
                'ca2_marks': result.ca2_marks,
                'home_work_marks': result.home_work_marks,
                'activity_marks': result.activity_marks,
                'exam_marks': result.exam_marks,
                'total_marks': result.ca1_marks + result.ca2_marks + result.home_work_marks + result.activity_marks + result.exam_marks,
                'grade': result.grade,
   
            })

        results_data.append({
            'student': student,
            'results': student_results,
            'total_score': total_score,
            'average_score': average_score,
            'overall_grade': overall_grade,
        })

    return render(request, 'src/display_class_results_summary.html', {
        'session': session,
        'term': term,
        'school_class': school_class,
        'results_data': results_data,
        'school_config': school_config,
    })



@login_required(login_url='login')
def export_results_to_excel(request, school_class_id, term_id, session_id):
    # Fetch data (same logic as in your original view)
    school_class = SchoolClass.objects.get(id=school_class_id)
    term = Term.objects.get(id=term_id)
    session = Session.objects.get(id=session_id)
    results_data = []  # Fetch this using your existing logic

    # Create a workbook and worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"{school_class.name} Results"

    # Add headers
    sheet.append(["Class", school_class.name])
    sheet.append(["Term", term.name, "Session", session.name])
    sheet.append([])  # Blank row

    # Create subject headers
    headers = ["Names/Subjects"]
    if results_data:
        for result in results_data[0]['results']:
            subject = result['subject']['name']
            headers.extend([f"{subject} - 1st C.A", f"{subject} - 2nd C.A", f"{subject} - Exams", f"{subject} - Total", f"{subject} - Grade"])

    sheet.append(headers)

    # Add student results
    for data in results_data:
        row = [data['student']['full_name']]
        for result in data['results']:
            row.extend([
                result['ca1_marks'],
                result['ca2_marks'],
                result['exam_marks'],
                result['total_marks'],
                result['grade'],
            ])
        sheet.append(row)

    # Adjust column alignment and widths
    for col in sheet.columns:
        for cell in col:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        col_width = max(len(str(cell.value or "")) for cell in col) + 2
        sheet.column_dimensions[col[0].column_letter].width = col_width

    # Generate response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename="{school_class.name}_results.xlsx"'
    workbook.save(response)

    return response



@login_required(login_url='login')
def result_entry_tahfeez(request):
    subjects = Subject.objects.all()
    classes = SchoolClass.objects.all()
    sessions = Session.objects.all()
    terms = Term.objects.all()

    if request.method == 'POST':
        term_id = request.POST.get('term')
        session_id = request.POST.get('session')
        class_id = request.POST.get('class_assigned')
        subject_id = request.POST.get('subject')

        if term_id and session_id and class_id and subject_id:
            term = get_object_or_404(Term, pk=term_id)
            session = get_object_or_404(Session, pk=session_id)
            school_class = get_object_or_404(SchoolClass, pk=class_id)
            subject = get_object_or_404(Subject, pk=subject_id)
            students = Student.objects.filter(enrolled_class=school_class)

            if 'save_results' in request.POST:
                for student in students:
                    # Ensure that marks are integers, defaulting to 0 if invalid or missing
                    marks = request.POST.get(f'marks_{student.id}', '0')


                    # Use int() conversion directly and handle exceptions
                    try:
                        marks = int(marks)
                    except ValueError:
                        marks = 0

                   

                    # Check if result already exists
                    if TahfeezResult.objects.filter(
                        student=student,
                        subject=subject,
                        class_assigned=school_class,
                        session=session,
                        term=term
                    ).exists():
                        messages.warning(request, f"Result for {student.first_name} {student.last_name} already exists.")
                    else:
                        # Create a new result if it doesn't exist
                        TahfeezResult.objects.create(
                            student=student,
                            subject=subject,
                            class_assigned=school_class,
                            session=session,
                            term=term,
                            marks=marks,

                        )

                messages.success(request, "Results processed. Check warnings for existing records.")
                return redirect('result_entry_tahfeez')

            return render(request, 'src/result_entry_tahfeez.html', {
                'subjects': subjects,
                'classes': classes,
                'sessions': sessions,
                'terms': terms,
                'students': students,
                'selected_term': term,
                'selected_session': session,
                'selected_class': school_class,
                'selected_subject': subject,
            })

    return render(request, 'src/result_entry_tahfeez.html', {
        'subjects': subjects,
        'classes': classes,
        'sessions': sessions,
        'terms': terms,
    })


@login_required(login_url='login')
def result_update_tahfeez(request):
    subjects = Subject.objects.all()
    classes = SchoolClass.objects.all()
    sessions = Session.objects.all()
    terms = Term.objects.all()

    if request.method == 'POST':
        term_id = request.POST.get('term')
        session_id = request.POST.get('session')
        class_id = request.POST.get('class_assigned')
        subject_id = request.POST.get('subject')

        if term_id and session_id and class_id and subject_id:
            term = get_object_or_404(Term, pk=term_id)
            session = get_object_or_404(Session, pk=session_id)
            school_class = get_object_or_404(SchoolClass, pk=class_id)
            subject = get_object_or_404(Subject, pk=subject_id)
            students = Student.objects.filter(enrolled_class=school_class)

            # Load existing results for the selected session, term, class, and subject
            results = TahfeezResult.objects.filter(
                session=session, term=term, class_assigned=school_class, subject=subject
            )

            # If the save button is clicked, update the results
            if 'save_results' in request.POST:
                for result in results:
                    # Ensure that marks are integers, defaulting to 0 if invalid or missing
                    marks = request.POST.get(f'marks_{result.student.id}', '0')
  

                    # Use int() conversion directly and handle exceptions
                    try:
                        result.marks = int(marks)
                    except ValueError:
                        result.ca1_marks = 0

                    

                    result.save()

                messages.success(request, "Results successfully updated.")
                return redirect('result_update_tahfeez')

            return render(request, 'src/result_update_tahfeez.html', {
                'subjects': subjects,
                'classes': classes,
                'sessions': sessions,
                'terms': terms,
                'students': students,
                'results': results,
                'selected_term': term,
                'selected_session': session,
                'selected_class': school_class,
                'selected_subject': subject,
            })

    return render(request, 'src/result_update_tahfeez.html', {
        'subjects': subjects,
        'classes': classes,
        'sessions': sessions,
        'terms': terms,
    })



@login_required(login_url='login')
def select_class_for_result_tahfeez(request):
    classes = SchoolClass.objects.all()
    sessions = Session.objects.all()
    terms = Term.objects.all()

    if request.method == 'POST':
        session_id = request.POST.get('session')
        term_id = request.POST.get('term')
        class_id = request.POST.get('class_assigned')

        if session_id and term_id and class_id:
            return redirect('display_class_results_tahfeez', session_id=session_id, term_id=term_id, class_id=class_id)

    return render(request, 'src/select_class_for_result.html', {
        'classes': classes,
        'sessions': sessions,
        'terms': terms,
    })




@login_required(login_url='login')
def display_class_results_tahfeez(request, session_id, term_id, class_id):
    session = get_object_or_404(Session, pk=session_id)
    term = get_object_or_404(Term, pk=term_id)
    school_class = get_object_or_404(SchoolClass, pk=class_id)
    students = Student.objects.filter(enrolled_class=school_class)
    school_config = SchoolConfig.objects.last()

    results_data = []

    for student in students:
        # Fetch all results for the student
        student_results = TahfeezResult.objects.filter(
            session=session,
            term=term,
            class_assigned=school_class,
            student=student
        )

        # Calculate total marks and grade
        total_marks = student_results.aggregate(total=Sum('marks'))['total'] or 0
        num_subjects = student_results.count()
        average_marks = total_marks / num_subjects if num_subjects > 0 else 0

        if student_results:
            overall_grade = student_results[0].grade
            class_position = student_results[0].class_position
        else:
            overall_grade = "Invalid Marks"
            class_position = "N/A"

        # Comments based on average score
        if 76 <= total_marks <= 100:
            english_comment = "AN EXCELLENT PERFORMANCE, KEEP IT UP."
            arabic_comment_male = "فاز بتقدير ممتاز ويرجى له التفوق في الفترات القادمة"
            arabic_comment_female = "فازت بتقدير ممتاز ويرجى لها التفوق في الفترات القادمة"
        elif 70 <= total_marks <= 76:
            english_comment = "AN EXCELLENT PERFORMANCE, KEEP IT UP."
            arabic_comment_male = "فاز بتقدير ممتاز ويرجى له التفوق في الفترات القادمة"
            arabic_comment_female = "فازت بتقدير ممتاز ويرجى لها التفوق في الفترات القادمة"
        elif 65 <= total_marks <= 70:
            english_comment = "AN EXCELLENT PERFORMANCE, KEEP IT UP."
            arabic_comment_male = "فاز بتقدير ممتاز ويرجى له التفوق في الفترات القادمة"
            arabic_comment_female = "فازت بتقدير ممتاز ويرجى لها التفوق في الفترات القادمة"
        elif 60 <= total_marks <= 65:
            english_comment = "A VERY GOOD RESULT, PUT IN MORE EFFORT."
            arabic_comment_male = "فاز بتقدير جيد جدا ويرجى له التقدم في الفترة المقبلة"
            arabic_comment_female = "فازت بتقدير جيد جدا ويرجى لها التقدم في الفترة المقبلة"
        elif 55 <= total_marks <= 60:
            english_comment = "A VERY GOOD RESULT, PUT IN MORE EFFORT."
            arabic_comment_male = "فاز بتقدير جيد جدا ويرجى له التقدم في الفترة المقبلة"
            arabic_comment_female = "فازت بتقدير جيد جدا ويرجى لها التقدم في الفترة المقبلة"
        elif 50 <= total_marks <= 55:
            english_comment = "A VERY GOOD RESULT, PUT IN MORE EFFORT."
            arabic_comment_male = "فاز بتقدير جيد جدا ويرجى له التقدم في الفترة المقبلة"
            arabic_comment_female = "فازت بتقدير جيد جدا ويرجى لها التقدم في الفترة المقبلة"
        elif 46 <= total_marks <= 50:
            english_comment = "A GOOD RESULT, TRY HARDER NEXT TERM."
            arabic_comment_male = "فاز بتقدير جيد ويرجى له الجهد الكبير في الفترة المقبلة"
            arabic_comment_female = "فازت بتقدير جيد ويرجى لها الجهد الكبير في الفترة المقبلة"
        elif 43 <= total_marks <= 46:
            english_comment = "A GOOD RESULT, TRY HARDER NEXT TERM."
            arabic_comment_male = "فاز بتقدير جيد ويرجى له الجهد الكبير في الفترة المقبلة"
            arabic_comment_female = "فازت بتقدير جيد ويرجى لها الجهد الكبير في الفترة المقبلة"
        elif 39 <= total_marks <= 43:
            english_comment = "A GOOD RESULT, TRY HARDER NEXT TERM."
            arabic_comment_male = "فاز بتقدير جيد ويرجى له الجهد الكبير في الفترة المقبلة"
            arabic_comment_female = "فازت بتقدير جيد ويرجى لها الجهد الكبير في الفترة المقبلة"
        elif 0 <= total_marks <= 39:
            english_comment = "A SATISFACTORY RESULT, TRY TO IMPROVE NEXT TERM."
            arabic_comment_male = "تقدير ضعيف،يرجى منه التقدم"
            arabic_comment_female = "تقدير ضعيف، يرجى منها التقدم"
        else:
            english_comment = "Invalid score."
            arabic_comment_male = arabic_comment_female = "درجة غير صالحة."

        # Determine gender-specific Arabic comment
        if student.gender == "Male":
            comments = f"{english_comment}\n{arabic_comment_male}"
        elif student.gender == "Female":
            comments = f"{english_comment}\n{arabic_comment_female}"
        else:
            comments = f"{english_comment}\nUnknown gender for Arabic comment."

        # Add results data for this student
        results_data.append({
            'student': student,
            'results': [
                {
                    'subject': result.subject,
                    'marks': result.marks,
                    'grade': result.grade
                }
                for result in student_results
            ],
            'total_marks': total_marks,
            'average_marks': average_marks,
            'overall_grade': overall_grade,
            'class_position': class_position,
            'comments': comments
        })

    return render(request, 'src/display_class_results_tahfeez.html', {
        'session': session,
        'term': term,
        'school_class': school_class,
        'results_data': results_data,
        'school_config': school_config,
    })



@login_required(login_url='login')
def delete_result(request):
    classes = SchoolClass.objects.all()
    sessions = Session.objects.all()
    terms = Term.objects.all()
    students = Student.objects.all()
    subjects = Subject.objects.all()

    if request.method == 'POST':
        # Collect filter parameters from the form
        session_id = request.POST.get('session')
        term_id = request.POST.get('term')
        class_id = request.POST.get('class_assigned')
        subject_id = request.POST.get('subjects')  # Fixed name
        student_id = request.POST.get('student')

        # Start building the query dynamically using Q objects
        query = Q()
        if session_id:
            query &= Q(session_id=session_id)
        if term_id:
            query &= Q(term_id=term_id)
        if class_id:
            query &= Q(class_assigned_id=class_id)
        if subject_id:
            query &= Q(subject_id=subject_id)
        if student_id:
            query &= Q(student_id=student_id)

        # Check if any filters were applied
        if query:
            deleted_count, _ = Result.objects.filter(query).delete()
            if deleted_count > 0:
                messages.success(request, f"{deleted_count} result(s) successfully deleted.")
            else:
                messages.info(request, "No results found for the given criteria.")
        else:
            messages.error(request, "Please select at least one filter to delete results.")

        return redirect('delete_results')  # Redirect to avoid resubmission

    return render(request, 'src/delete_result.html', {
        'classes': classes,
        'sessions': sessions,
        'terms': terms,
        'subjects': subjects,
        'students': students,
    })


@login_required(login_url='login')
def delete_result_tahfeez(request):
    classes = SchoolClass.objects.all()
    sessions = Session.objects.all()
    terms = Term.objects.all()
    students = Student.objects.all()
    subjects = Subject.objects.all()

    if request.method == 'POST':
        # Collect filter parameters from the form
        session_id = request.POST.get('session')
        term_id = request.POST.get('term')
        class_id = request.POST.get('class_assigned')
        subject_id = request.POST.get('subjects')  # Fixed name
        student_id = request.POST.get('student')

        # Start building the query dynamically using Q objects
        query = Q()
        if session_id:
            query &= Q(session_id=session_id)
        if term_id:
            query &= Q(term_id=term_id)
        if class_id:
            query &= Q(class_assigned_id=class_id)
        if subject_id:
            query &= Q(subject_id=subject_id)
        if student_id:
            query &= Q(student_id=student_id)

        # Check if any filters were applied
        if query:
            deleted_count, _ = TahfeezResult.objects.filter(query).delete()
            if deleted_count > 0:
                messages.success(request, f"{deleted_count} result(s) successfully deleted.")
            else:
                messages.info(request, "No results found for the given criteria.")
        else:
            messages.error(request, "Please select at least one filter to delete results.")

        return redirect('delete_result_tahfeez')  # Redirect to avoid resubmission

    return render(request, 'src/delete_result.html', {
        'classes': classes,
        'sessions': sessions,
        'terms': terms,
        'subjects': subjects,
        'students': students,
    })

